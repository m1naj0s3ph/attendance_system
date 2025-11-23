from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session, jsonify
import sqlite3
import os
from datetime import datetime
from io import BytesIO
from urllib.parse import quote
import csv
import re

# ---------- Config ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "students.db")
EXCEL_PATH = os.path.join(BASE_DIR, "students.xlsx")
QR_DIR = os.path.join(BASE_DIR, "static", "qr_codes")
SUMMARY_DIR = os.path.join(BASE_DIR, "summary_of_the_day")
MONTHLY_DIR = os.path.join(BASE_DIR, "monthly_reports")

# إنشاء المجلدات إذا لم تكن موجودة
os.makedirs(QR_DIR, exist_ok=True)
os.makedirs(SUMMARY_DIR, exist_ok=True)
os.makedirs(MONTHLY_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = "attendance-system-secret-key-2024-pythonanywhere"

# ---------- نظام المستخدمين ----------
USERS = {
    "admin": "Admin123!",
    "teacher": "teacher123"
}

USER_PERMISSIONS = {
    "admin": ["all"],
    "teacher": ["view", "scan", "add_record", "daily_report", "bulk_grades"]
}

@app.before_request
def require_login():
    public_pages = ['login', 'static', 'logout', 'remote_scanner']
    if request.endpoint and not any(request.endpoint == page or request.endpoint.startswith('static') for page in public_pages):
        if 'username' not in session:
            return redirect(url_for('login'))

def check_permission(required_permission):
    if 'username' not in session:
        return False
    username = session['username']
    if username not in USER_PERMISSIONS:
        return False
    permissions = USER_PERMISSIONS[username]
    return 'all' in permissions or required_permission in permissions

# ---------- Helpers ----------
def get_pythonanywhere_url():
    """الحصول على رابط PythonAnywhere"""
    username = os.environ.get('PYTHONANYWHERE_USERNAME', 'yourusername')
    return f"{username}.pythonanywhere.com"

PC_IP = get_pythonanywhere_url()

def open_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def today_str():
    return datetime.now().strftime("%Y-%m-%d")

def current_month_str():
    return datetime.now().strftime("%Y-%m")

def weekday_english_to_arabic(day_english):
    day_map = {
        "sunday": "الأحد",
        "monday": "الإثنين",
        "tuesday": "الثلاثاء",
        "wednesday": "الأربعاء",
        "thursday": "الخميس",
        "friday": "الجمعة",
        "saturday": "السبت"
    }
    return day_map.get(day_english.lower(), day_english)

# ---------- Database Initialization ----------
def init_tables():
    """إنشاء جميع الجداول إذا لم تكن موجودة"""
    conn = open_db()

    conn.execute("""
    CREATE TABLE IF NOT EXISTS students (
        id TEXT PRIMARY KEY,
        student_name TEXT,
        parent_number TEXT,
        payment_amount REAL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS classes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        day_of_week TEXT,
        start_time TEXT,
        end_time TEXT,
        FOREIGN KEY (student_id) REFERENCES students (id)
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        class_id INTEGER,
        exam_grade TEXT,
        homework_status TEXT,
        status TEXT,
        paid TEXT,
        date TEXT,
        FOREIGN KEY (student_id) REFERENCES students (id),
        FOREIGN KEY (class_id) REFERENCES classes (id)
    )
    """)

    conn.commit()
    conn.close()
    print("✅ تم إنشاء/التأكد من جميع الجداول في قاعدة البيانات")

def init_db_from_excel():
    """تهيئة قاعدة البيانات من ملف Excel"""
    if not os.path.exists(EXCEL_PATH):
        print(f"⚠️  تحذير: ملف {EXCEL_PATH} غير موجود")
        print("📝 الرجاء إنشاء ملف Excel بالهيكل التالي:")
        print("   الأعمدة: id, student_name, parent_number, payment_amount, day_of_week")
        return

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(EXCEL_PATH)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        conn = open_db()

        students_added = 0
        classes_added = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            student_data = dict(zip(headers, row))

            conn.execute("""
                INSERT OR REPLACE INTO students (id, student_name, parent_number, payment_amount)
                VALUES (?, ?, ?, ?)
            """, (
                str(student_data.get('id', '')),
                str(student_data.get('student_name', '')),
                str(student_data.get('parent_number', '')),
                float(student_data.get('payment_amount', 0))
            ))
            students_added += 1

            day_of_week = student_data.get('day_of_week')

            if day_of_week:
                day_of_week = str(day_of_week).strip().lower()

                conn.execute("""
                    INSERT INTO classes (student_id, day_of_week, start_time, end_time)
                    VALUES (?, ?, ?, ?)
                """, (str(student_data['id']), day_of_week, "09:00", "10:00"))
                classes_added += 1

        conn.commit()
        conn.close()

        qr_created = 0
        workbook = load_workbook(EXCEL_PATH)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                student_id = str(row[0])
                if generate_qr(student_id):
                    qr_created += 1

        print(f"✅ تم تحميل بيانات {students_added} طالب و {classes_added} حصة من ملف Excel")
        print(f"✅ تم إنشاء {qr_created} رمز QR")

    except Exception as e:
        print(f"❌ خطأ في تحميل بيانات Excel: {e}")

# ---------- Class Management ----------
def get_student_classes(student_id):
    """جلب جميع حصص الطالب"""
    conn = open_db()
    cursor = conn.execute("SELECT * FROM classes WHERE student_id=?", (student_id,))
    classes = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return classes

def get_current_class(student_id):
    """الحصول على الحصة الحالية للطالب بناءً على اليوم فقط"""
    conn = open_db()
    current_day = datetime.now().strftime("%A").lower()

    cursor = conn.execute("""
        SELECT * FROM classes
        WHERE student_id=? AND day_of_week=?
    """, (student_id, current_day))

    class_row = cursor.fetchone()
    if class_row:
        conn.close()
        return dict(class_row)

    conn.close()
    return None

def is_class_today(student_id):
    """التحقق إذا كان الطالب لديه حصة اليوم"""
    return get_current_class(student_id) is not None

def get_today_classes(student_id):
    """جلب حصص الطالب لليوم الحالي"""
    conn = open_db()
    current_day = datetime.now().strftime("%A").lower()
    cursor = conn.execute("""
        SELECT * FROM classes
        WHERE student_id=? AND day_of_week=?
    """, (student_id, current_day))

    classes = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return classes

def get_weekly_classes(student_id):
    """جلب جميع حصص الطالب للأسبوع"""
    conn = open_db()
    cursor = conn.execute("SELECT * FROM classes WHERE student_id=? ORDER BY day_of_week", (student_id,))
    classes = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return classes

# ---------- QR generation ----------
def generate_qr(student_id):
    """إنشاء QR Code"""
    try:
        import qrcode
        
        # إنشاء QR code محلياً
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        
        # استخدام رابط PythonAnywhere
        link = f"https://{PC_IP}/student/{student_id}"
        qr.add_data(link)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        path = os.path.join(QR_DIR, f"{student_id}.png")
        img.save(path)
        
        print(f"✅ تم إنشاء QR للطالب {student_id}")
        return path
    except Exception as e:
        print(f"❌ خطأ في إنشاء QR للطالب {student_id}: {e}")
        return None

# ---------- Routes للتحكم عن بعد ----------
@app.route("/remote_scanner")
def remote_scanner():
    """صفحة الماسح الضوئي للتليفون"""
    return render_template("remote_scanner.html", ip=PC_IP)

# ---------- WhatsApp Messages ----------
def generate_whatsapp_link(student_data):
    """إنشاء رابط واتساب ويب لرسالة الغياب"""
    try:
        student_name = student_data['student_name']
        parent_number = student_data['parent_number']
        date = today_str()

        message = f"""
تنبيه غياب
عزيزي ولي الأمر،
الطالب/ة {student_name} لم يحضر الحصة اليوم {date}.

يرجى التواصل مع الإدارة للاستفسار عن سبب الغياب.

مع تحيات،
الإدارة
        """.strip()

        phone_number = ''.join(filter(str.isdigit, str(parent_number)))

        if not phone_number:
            print(f"❌ رقم ولي الأمر غير صالح للطالب {student_name}")
            return None

        if not phone_number.startswith('+'):
            phone_number = '+2' + phone_number

        encoded_message = quote(message)
        whatsapp_link = f"https://web.whatsapp.com/send?phone={phone_number}&text={encoded_message}"

        return whatsapp_link

    except Exception as e:
        print(f"❌ خطأ في إنشاء رابط واتساب: {e}")
        return None

def generate_detailed_monthly_report_message(student_id, month_str=None):
    """إنشاء رسالة واتساب مفصلة للتقرير الشهري"""
    if month_str is None:
        month_str = current_month_str()

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()
    if not student_row:
        conn.close()
        return None

    student_data = dict(student_row)

    cursor = conn.execute(
        "SELECT * FROM history WHERE student_id=? AND date LIKE ? ORDER BY date ASC",
        (student_id, f"{month_str}-%")
    )
    history_rows = [dict(row) for row in cursor.fetchall()]
    conn.close()

    total_classes = len(history_rows)
    present_count = len([h for h in history_rows if h['status'] == 'Present'])
    absent_count = len([h for h in history_rows if h['status'] == 'Absent'])

    if total_classes > 0:
        attendance_rate = (present_count / total_classes) * 100
    else:
        attendance_rate = 0

    payment_amount = student_data.get('payment_amount', 0)
    paid_sessions = [h for h in history_rows if h['paid'] == 'Yes']
    paid_amount = len(paid_sessions) * float(payment_amount)

    exam_grades = [h['exam_grade'] for h in history_rows if h['exam_grade'] != '-']
    if exam_grades:
        try:
            grades_numeric = []
            for grade in exam_grades:
                try:
                    grades_numeric.append(float(grade))
                except:
                    continue

            if grades_numeric:
                avg_grade = sum(grades_numeric) / len(grades_numeric)
                avg_grade = f"{avg_grade:.1f}"
            else:
                avg_grade = "لا توجد درجات"
        except:
            avg_grade = "لا توجد درجات"
    else:
        avg_grade = "لا توجد درجات"

    homework_done = len([h for h in history_rows if h['homework_status'] == 'اتعمل'])
    homework_not_done = len([h for h in history_rows if h['homework_status'] == 'متعملش'])

    message = f"""
📊 **التقرير الشهري للطالب/ة {student_data['student_name']}**
🗓️ **الشهر:** {month_str}

**📈 الإحصائيات العامة:**
• إجمالي الحصص: {total_classes}
• عدد الحضور: {present_count}
• عدد الغياب: {absent_count}
• معدل الحضور: {attendance_rate:.1f}%

**💰 الجانب المالي:**
• المبلغ المدفوع: {paid_amount:.2f} ج.م
• قيمة الحصة: {payment_amount} ج.م

**📚 الأداء الأكاديمي:**
• متوسط الدرجات: {avg_grade}
• الواجبات المنجزة: {homework_done}
• الواجبات غير المنجزة: {homework_not_done}

**📅 تفاصيل الحصص:**
"""

    for session in history_rows:
        status_icon = "✅" if session['status'] == 'Present' else "❌"
        homework_icon = "✅" if session['homework_status'] == 'اتعمل' else "❌" if session['homework_status'] == 'متعملش' else "➖"
        paid_icon = "💰" if session['paid'] == 'Yes' else "❌"

        message += f"\n{status_icon} {session['date']}: "
        message += f"امتحان({session['exam_grade']}) "
        message += f"واجب{homework_icon} "
        message += f"دفع{paid_icon}"

    message += "\n\nمع تحيات الإدارة 🏫"
    return message

def generate_present_student_message(student_data):
    """إنشاء رسالة واتساب للطالب الحاضر"""
    try:
        student_name = student_data['student_name']
        parent_number = student_data['parent_number']
        date = today_str()
        exam_grade = student_data.get('exam_grade', '-')
        homework_status = student_data.get('homework_status', '-')

        homework_text = ""
        if homework_status == 'اتعمل':
            homework_text = "تم إنهاء الواجب بنجاح ✅"
        elif homework_status == 'متعملش':
            homework_text = "لم يتم إنهاء الواجب ❌"
        else:
            homework_text = "لا يوجد واجب لهذا اليوم"

        grade_text = ""
        if exam_grade != '-' and exam_grade != '':
            grade_text = f"• درجة الامتحان: {exam_grade}"
        else:
            grade_text = "• لم يتم تسجيل درجة امتحان لهذا اليوم"

        message = f"""
تقرير الحصة اليومية
عزيزي ولي الأمر،

الطالب/ة {student_name} قد حضر الحصة اليوم {date} بنجاح.

تفاصيل اليوم:
• الحالة: حاضر ✅
{grade_text}
• الواجب: {homework_text}

نشكر متابعتكم ودعمكم المستمر لتحسين أداء الطالب/ة.

مع تحيات،
الإدارة
        """.strip()

        phone_number = ''.join(filter(str.isdigit, str(parent_number)))

        if not phone_number:
            print(f"❌ رقم ولي الأمر غير صالح للطالب {student_name}")
            return None

        if not phone_number.startswith('+'):
            phone_number = '+2' + phone_number

        encoded_message = quote(message)
        whatsapp_link = f"https://web.whatsapp.com/send?phone={phone_number}&text={encoded_message}"

        return whatsapp_link

    except Exception as e:
        print(f"❌ خطأ في إنشاء رابط واتساب للطالب الحاضر: {e}")
        return None

def check_and_generate_whatsapp_links():
    """التحقق من الغياب وإنشاء روابط واتساب"""
    print("🔍 جاري فحص الغياب وإنشاء روابط واتساب...")

    conn = open_db()
    date = today_str()

    cursor = conn.execute("""
        SELECT DISTINCT s.*
        FROM students s
        JOIN classes c ON s.id = c.student_id
        LEFT JOIN history h ON s.id = h.student_id AND h.date = ?
        WHERE (h.status IS NULL OR h.status = 'Absent')
        AND c.day_of_week = ?
    """, (date, datetime.now().strftime("%A").lower()))

    absent_students = [dict(row) for row in cursor.fetchall()]
    conn.close()

    if not absent_students:
        print("✅ لا يوجد طلاب غائبين اليوم")
        return []

    print(f"📋 عدد الطلاب الغائبين: {len(absent_students)}")

    whatsapp_links = []
    for student in absent_students:
        link = generate_whatsapp_link(student)
        if link:
            whatsapp_links.append({
                'student_name': student['student_name'],
                'parent_number': student['parent_number'],
                'whatsapp_link': link
            })

    print(f"✅ تم إنشاء {len(whatsapp_links)} رابط واتساب")
    return whatsapp_links

# ---------- Auto-mark absent ----------
def mark_absent_for_today():
    """وضع علامة غياب للطلاب الذين لديهم حصة اليوم ولم يسجلوا حضور"""
    conn = open_db()
    date = today_str()
    current_day = datetime.now().strftime("%A").lower()

    cursor = conn.execute("""
        SELECT s.id as student_id, c.id as class_id
        FROM students s
        JOIN classes c ON s.id = c.student_id
        WHERE c.day_of_week = ?
        AND NOT EXISTS (
            SELECT 1 FROM history h 
            WHERE h.student_id = s.id AND h.date = ?
        )
    """, (current_day, date))

    students_to_mark_absent = [dict(row) for row in cursor.fetchall()]
    
    marked_count = 0
    for row in students_to_mark_absent:
        check_cursor = conn.execute("""
            SELECT 1 FROM history 
            WHERE student_id=? AND class_id=? AND date=?
        """, (row['student_id'], row['class_id'], date))
        
        if not check_cursor.fetchone():
            conn.execute("""
                INSERT INTO history (student_id, class_id, exam_grade, homework_status, status, paid, date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (row['student_id'], row['class_id'], "-", "-", "Absent", "No", date))
            marked_count += 1

    conn.commit()
    conn.close()
    
    if marked_count > 0:
        print(f"✅ تم تعيين {marked_count} طالب كغائبين لليوم")
    else:
        print("✅ لا يوجد طلاب يحتاجون وضع غياب")

# ---------- Save daily summary ----------
def save_daily_summary():
    """حفظ التقرير اليومي كملف CSV"""
    conn = open_db()
    date = today_str()

    cursor = conn.execute("SELECT * FROM history WHERE date=?", (date,))
    history_today = [dict(row) for row in cursor.fetchall()]

    cursor = conn.execute("SELECT * FROM students")
    students = [dict(row) for row in cursor.fetchall()]

    conn.close()

    merged = []
    for st in students:
        rec = next((h for h in history_today if h["student_id"] == st["id"]), None)
        merged.append({
            "id": st["id"],
            "student_name": st.get("student_name"),
            "parent_number": st.get("parent_number"),
            "exam_grade": rec["exam_grade"] if rec else "-",
            "homework_status": rec["homework_status"] if rec else "-",
            "status": rec["status"] if rec else "Absent",
            "paid": rec["paid"] if rec else "No",
            "payment_amount": st.get("payment_amount", 0)
        })

    filename = f"{date}.csv"
    filepath = os.path.join(SUMMARY_DIR, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['id', 'student_name', 'parent_number', 'exam_grade', 'homework_status', 'status', 'paid', 'payment_amount']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(merged)

    return filepath

# ---------- Monthly report generator ----------
def generate_monthly_report_file(student_id, month_str=None):
    """إنشاء تقرير شهري كملف CSV"""
    if month_str is None:
        month_str = current_month_str()

    conn = open_db()

    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()
    if not student_row:
        conn.close()
        return None

    student = dict(student_row)
    
    student_name = student.get("student_name", "unknown")
    student_name_safe = re.sub(r'[^\w\s\u0600-\u06FF]', '', student_name)
    student_name_safe = re.sub(r'\s+', '_', student_name_safe.strip())
    
    if not student_name_safe:
        student_name_safe = f"student_{student_id}"
    
    filename = f"{student_name_safe}_{month_str}.csv"
    filepath = os.path.join(MONTHLY_DIR, filename)

    try:
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            csvfile.write(f"Student ID,{student_id}\n")
            csvfile.write(f"Student Name,{student.get('student_name','')}\n")
            csvfile.write(f"Month,{month_str}\n\n")

            fieldnames = ['date', 'status', 'homework_status', 'exam_grade', 'paid']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            cursor = conn.execute("SELECT * FROM history WHERE student_id=? AND date LIKE ? ORDER BY date ASC", (student_id, f"{month_str}-%"))
            history_rows = [dict(row) for row in cursor.fetchall()]
            
            if not history_rows:
                writer.writerow({
                    'date': 'لا توجد بيانات', 'status': '-',
                    'homework_status': '-', 'exam_grade': '-', 'paid': '-'
                })
            else:
                for session in history_rows:
                    writer.writerow({
                        'date': session['date'],
                        'status': session['status'],
                        'homework_status': session['homework_status'],
                        'exam_grade': session['exam_grade'],
                        'paid': session['paid']
                    })

        conn.close()
        return filepath
        
    except Exception as e:
        print(f"❌ خطأ في إنشاء التقرير الشهري: {e}")
        conn.close()
        return None

def calculate_monthly_stats():
    """حساب إحصائيات الشهر للطلاب"""
    month_str = current_month_str()
    conn = open_db()

    cursor = conn.execute("SELECT * FROM students")
    students = [dict(row) for row in cursor.fetchall()]

    monthly_stats = []

    for student in students:
        student_id = student["id"]
        student_name = student["student_name"]
        parent_number = student["parent_number"]
        payment_amount = student.get("payment_amount", 0)

        cursor = conn.execute(
            "SELECT * FROM history WHERE student_id=? AND date LIKE ? ORDER BY date ASC",
            (student_id, f"{month_str}-%")
        )
        history_rows = [dict(row) for row in cursor.fetchall()]

        total_classes = len(history_rows)
        present_count = len([h for h in history_rows if h['status'] == 'Present'])
        absent_count = len([h for h in history_rows if h['status'] == 'Absent'])

        if total_classes > 0:
            attendance_rate = (present_count / total_classes) * 100
        else:
            attendance_rate = 0

        paid_amount = 0
        paid_sessions = [h for h in history_rows if h['paid'] == 'Yes']
        if paid_sessions:
            paid_amount = len(paid_sessions) * float(payment_amount)

        cursor = conn.execute("SELECT DISTINCT day_of_week FROM classes WHERE student_id=?", (student_id,))
        class_days_rows = [dict(row) for row in cursor.fetchall()]
        class_days = [weekday_english_to_arabic(row['day_of_week']) for row in class_days_rows]
        class_days_str = ", ".join(class_days) if class_days else "لا توجد حصص"

        monthly_stats.append({
            "id": student_id,
            "student_name": student_name,
            "parent_number": parent_number,
            "class_days": class_days_str,
            "total_classes": total_classes,
            "present_count": present_count,
            "absent_count": absent_count,
            "attendance_rate": attendance_rate,
            "paid_amount": paid_amount,
            "payment_amount": payment_amount
        })

    conn.close()
    return monthly_stats

# ---------- Student Management ----------
@app.route("/add_student", methods=["GET", "POST"])
def add_student():
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    if request.method == "POST":
        try:
            student_id = request.form.get("student_id", "").strip()
            student_name = request.form.get("student_name", "").strip()
            parent_number = request.form.get("parent_number", "").strip()
            payment_amount = request.form.get("payment_amount", "0").strip()

            if not student_id or not student_name:
                flash("يرجى إدخال رقم الطالب والاسم", "error")
                return render_template("add_student.html")

            try:
                payment_amount = float(payment_amount)
            except:
                payment_amount = 0.0

            conn = open_db()

            cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
            existing = cursor.fetchone()
            if existing:
                flash("رقم الطالب موجود مسبقاً", "error")
                conn.close()
                return render_template("add_student.html")

            conn.execute("""
                INSERT INTO students (id, student_name, parent_number, payment_amount)
                VALUES (?, ?, ?, ?)
            """, (student_id, student_name, parent_number, payment_amount))

            days_of_week = request.form.getlist("day_of_week[]")

            classes_added = 0
            for day in days_of_week:
                day = day.strip().lower()
                if day:
                    conn.execute("""
                        INSERT INTO classes (student_id, day_of_week, start_time, end_time)
                        VALUES (?, ?, ?, ?)
                    """, (student_id, day, "09:00", "10:00"))
                    classes_added += 1

            conn.commit()
            conn.close()

            qr_result = generate_qr(student_id)

            if qr_result:
                print(f"✅ تم إنشاء QR للطالب {student_id}")
            else:
                print(f"⚠️  فشل إنشاء QR للطالب {student_id}")

            try:
                from openpyxl import load_workbook, Workbook
                
                if os.path.exists(EXCEL_PATH):
                    workbook = load_workbook(EXCEL_PATH)
                    sheet = workbook.active
                else:
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.append(['id', 'student_name', 'parent_number', 'payment_amount', 'day_of_week'])
                
                for day in days_of_week:
                    day = day.strip().lower()
                    if day:
                        sheet.append([student_id, student_name, parent_number, payment_amount, day])
                
                workbook.save(EXCEL_PATH)
                print(f"✅ تم إضافة الطالب {student_name} إلى ملف Excel")
            except Exception as e:
                print(f"⚠️  تحذير: لم يتم إضافة الطالب إلى ملف Excel: {e}")

            flash(f"تم إضافة الطالب {student_name} بنجاح وإنشاء QR code", "success")
            return redirect(url_for("manage_students"))

        except Exception as e:
            flash(f"خطأ في إضافة الطالب: {e}", "error")
            return render_template("add_student.html")

    return render_template("add_student.html")

@app.route("/manage_students")
def manage_students():
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students ORDER BY id")
    students = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template("manage_students.html", students=students, get_student_classes=get_student_classes)

@app.route("/delete_student/<student_id>")
def delete_student(student_id):
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    try:
        conn = open_db()
        conn.execute("DELETE FROM history WHERE student_id=?", (student_id,))
        conn.execute("DELETE FROM classes WHERE student_id=?", (student_id,))
        conn.execute("DELETE FROM students WHERE id=?", (student_id,))
        conn.commit()
        conn.close()

        qr_path = os.path.join(QR_DIR, f"{student_id}.png")
        if os.path.exists(qr_path):
            os.remove(qr_path)

        flash("تم حذف الطالب بنجاح", "success")
    except Exception as e:
        flash(f"خطأ في حذف الطالب: {e}", "error")

    return redirect(url_for("manage_students"))

# ---------- Class Management Routes ----------
@app.route("/add_class/<student_id>", methods=["POST"])
def add_class(student_id):
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    day_of_week = request.form.get("day_of_week", "").strip().lower()
    start_time = request.form.get("start_time", "").strip()
    end_time = request.form.get("end_time", "").strip()

    if not day_of_week or not start_time or not end_time:
        flash("يرجى ملء جميع الحقول", "error")
        return redirect(url_for("manage_students"))

    try:
        conn = open_db()
        conn.execute("""
            INSERT INTO classes (student_id, day_of_week, start_time, end_time)
            VALUES (?, ?, ?, ?)
        """, (student_id, day_of_week, start_time, end_time))
        conn.commit()
        conn.close()

        flash("تم إضافة الحصة بنجاح", "success")
    except Exception as e:
        flash(f"خطأ في إضافة الحصة: {e}", "error")

    return redirect(url_for("manage_students"))

@app.route("/delete_class/<class_id>")
def delete_class(class_id):
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    try:
        conn = open_db()
        conn.execute("DELETE FROM classes WHERE id=?", (class_id,))
        conn.commit()
        conn.close()

        flash("تم حذف الحصة بنجاح", "success")
    except Exception as e:
        flash(f"خطأ في حذف الحصة: {e}", "error")

    return redirect(url_for("manage_students"))

# ---------- Bulk Grades Management ----------
@app.route("/bulk_grades", methods=["GET", "POST"])
def bulk_grades():
    if not check_permission('bulk_grades'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    if request.method == "POST":
        student_id = request.form.get("student_id", "").strip()
        grade = request.form.get("grade", "").strip()
        hw_status = request.form.get("hw_status", "").strip()
        date = today_str()

        if not student_id:
            flash("يرجى إدخال رقم الطالب", "error")
            return redirect(url_for("bulk_grades"))

        conn = open_db()

        cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
        student_row = cursor.fetchone()
        if not student_row:
            flash("رقم الطالب غير موجود", "error")
            conn.close()
            return redirect(url_for("bulk_grades"))

        cursor = conn.execute("SELECT * FROM history WHERE student_id=? AND date=?", (student_id, date))
        existing_record = cursor.fetchone()

        if existing_record:
            conn.execute("""
                UPDATE history SET exam_grade=?, homework_status=?, status='Present'
                WHERE student_id=? AND date=?
            """, (grade or "-", hw_status or "-", student_id, date))
        else:
            today_classes = get_today_classes(student_id)
            if today_classes:
                class_id = today_classes[0]["id"]
            else:
                class_id = None

            conn.execute("""
                INSERT INTO history (student_id, class_id, exam_grade, homework_status, status, paid, date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (student_id, class_id, grade or "-", hw_status or "-", "Present", "Yes", date))

        conn.commit()
        conn.close()

        student_name = student_row['student_name']
        flash(f"تم تحديث بيانات الطالب {student_name} بنجاح", "success")
        return redirect(url_for("bulk_grades"))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students ORDER BY id")
    students = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return render_template("bulk_grades.html", students=students, get_student_classes=get_student_classes)

# ---------- Routes ----------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if username in USERS and USERS[username] == password:
            session['logged_in'] = True
            session['username'] = username
            flash(f"مرحباً {username}!", "success")
            return redirect(url_for('index'))
        else:
            flash("اسم المستخدم أو كلمة المرور غير صحيحة", "error")

    if 'username' in session:
        return redirect(url_for('index'))

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("تم تسجيل الخروج بنجاح", "success")
    return redirect(url_for('login'))

@app.route("/")
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template("display.html", ip=PC_IP, username=session.get('username'))

@app.route("/student/<student_id>")
def student_page(student_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()
    if not student_row:
        conn.close()
        return "لا يوجد طالب بهذا الكود", 404

    student = dict(student_row)
    date = today_str()

    current_class = get_current_class(student_id)
    today_classes = get_today_classes(student_id)
    weekly_classes = get_weekly_classes(student_id)

    if current_class:
        cursor = conn.execute("""
            SELECT * FROM history
            WHERE student_id=? AND class_id=? AND date=?
        """, (student_id, current_class["id"], date))

        row = cursor.fetchone()
        if row and row["status"] == "Absent":
            conn.execute("""
                UPDATE history SET status='Present', paid='Yes'
                WHERE student_id=? AND class_id=? AND date=?
            """, (student_id, current_class["id"], date))
            conn.commit()
        elif not row:
            conn.execute("""
                INSERT INTO history (student_id, class_id, exam_grade, homework_status, status, paid, date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (student_id, current_class["id"], "-", "-", "Present", "Yes", date))
            conn.commit()
    else:
        flash("⚠️ اليوم ليس يوم حصة للطالب، لم يتم تسجيل الحضور", "warning")

    cursor = conn.execute("""
        SELECT * FROM history
        WHERE student_id=? AND date=?
        ORDER BY id DESC LIMIT 1
    """, (student_id, date))

    latest_record = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return render_template("student.html",
                         student=student,
                         history=latest_record,
                         current_class=current_class,
                         today_classes=today_classes,
                         weekly_classes=weekly_classes,
                         today=date,
                         now=datetime.now(),
                         username=session.get('username'))

@app.route("/add_record/<student_id>", methods=["POST"])
def add_record(student_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    if not is_class_today(student_id):
        flash("⚠️ اليوم ليس يوم حصة للطالب، لا يمكن تعديل البيانات", "error")
        return redirect(url_for("student_page", student_id=student_id))

    grade = request.form.get("grade", "").strip()
    hw = request.form.get("hw", "").strip()
    date = today_str()

    conn = open_db()

    cursor = conn.execute("SELECT * FROM history WHERE student_id=? AND date=? ORDER BY id DESC LIMIT 1", (student_id, date))
    existing_record = cursor.fetchone()

    if existing_record:
        conn.execute("""
            UPDATE history SET exam_grade=?, homework_status=?, status='Present', paid='Yes'
            WHERE student_id=? AND date=?
        """, (grade or "-", hw or "-", student_id, date))
    else:
        today_classes = get_today_classes(student_id)
        if today_classes:
            class_id = today_classes[0]["id"]
        else:
            class_id = None

        conn.execute("""
            INSERT INTO history (student_id, class_id, exam_grade, homework_status, status, paid, date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (student_id, class_id, grade or "-", hw or "-", "Present", "Yes", date))

    conn.commit()
    conn.close()

    return redirect(url_for("index"))

@app.route("/direct_scan", methods=["POST"])
def direct_scan():
    if 'username' not in session:
        return redirect(url_for('login'))

    student_id = request.form.get("student_id", "").strip()
    if not student_id:
        return redirect(url_for("index"))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()
    if not student_row:
        conn.close()
        return redirect(url_for("index"))

    conn.close()
    return redirect(url_for("student_page", student_id=student_id))

@app.route("/admin")
def admin():
    if 'username' not in session:
        return redirect(url_for('login'))

    monthly_stats = calculate_monthly_stats()
    month_str = current_month_str()

    total_students = len(monthly_stats)
    total_paid = sum(stats['paid_amount'] for stats in monthly_stats)
    total_present = sum(stats['present_count'] for stats in monthly_stats)

    total_attendance = total_present + sum(stats['absent_count'] for stats in monthly_stats)
    if total_attendance > 0:
        overall_attendance = (total_present / total_attendance) * 100
    else:
        overall_attendance = 0

    return render_template("admin.html",
                         students=monthly_stats,
                         total_paid=total_paid,
                         total_students=total_students,
                         total_present=total_present,
                         overall_attendance=overall_attendance,
                         month=month_str,
                         ip=PC_IP,
                         username=session.get('username'))

@app.route("/daily_report")
def daily_report():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students")
    students = [dict(row) for row in cursor.fetchall()]
    date = today_str()

    cursor = conn.execute("SELECT * FROM history WHERE date=?", (date,))
    history_today = [dict(row) for row in cursor.fetchall()]
    conn.close()

    merged = []
    total_paid = 0.0
    for st in students:
        rec = next((h for h in history_today if h["student_id"] == st["id"]), None)
        status = rec["status"] if rec else "Absent"
        paid = rec["paid"] if rec else "No"
        exam = rec["exam_grade"] if rec else "-"
        hw = rec["homework_status"] if rec else "-"
        payment_amount = st.get("payment_amount") or 0

        merged.append({
            "id": st["id"],
            "student_name": st.get("student_name"),
            "parent_number": st.get("parent_number"),
            "status": status,
            "paid": paid,
            "exam_grade": exam,
            "homework_status": hw,
            "payment_amount": payment_amount
        })

        if paid == "Yes":
            try:
                total_paid += float(payment_amount)
            except:
                pass

    save_daily_summary()

    return render_template("daily_report.html", students=merged, total_paid=total_paid, date=date, ip=PC_IP, username=session.get('username'))

@app.route("/monthly_report/<student_id>")
def monthly_report(student_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    month = request.args.get("month", None)
    if month is None:
        month = current_month_str()
    path = generate_monthly_report_file(student_id, month)
    if not path or not os.path.exists(path):
        flash("❌ تعذر إنشاء التقرير الشهري", "error")
        return redirect(url_for("student_page", student_id=student_id))
    return send_file(path, as_attachment=True)

@app.route("/download_daily_summary")
def download_daily_summary():
    if 'username' not in session:
        return redirect(url_for('login'))

    filepath = save_daily_summary()
    return send_file(filepath, as_attachment=True)

@app.route("/download_all_reports")
def download_all_reports():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT id FROM students")
    student_ids = [row['id'] for row in cursor.fetchall()]
    conn.close()

    from zipfile import ZipFile
    import io

    zip_buffer = io.BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for student_id in student_ids:
            report_path = generate_monthly_report_file(student_id)
            if report_path and os.path.exists(report_path):
                zip_file.write(report_path, os.path.basename(report_path))

    zip_buffer.seek(0)
    return send_file(zip_buffer, download_name=f"monthly_reports_{current_month_str()}.zip", as_attachment=True)

@app.route("/reload_students")
def reload_students():
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    init_db_from_excel()
    flash("تم تحديث البيانات من ملف Excel بنجاح", "success")
    return redirect(url_for('admin'))

@app.route("/generate_all_qr")
def generate_all_qr():
    if not check_permission('all'):
        flash("غير مصرح لك بهذا الإجراء", "error")
        return redirect(url_for('index'))

    conn = open_db()
    cursor = conn.execute("SELECT id FROM students")
    student_ids = [row['id'] for row in cursor.fetchall()]
    conn.close()

    for student_id in student_ids:
        generate_qr(student_id)

    flash("تم إنشاء رموز QR لجميع الطلاب", "success")
    return redirect(url_for('admin'))

# ---------- Routes للرسائل النصية ----------
@app.route("/generate_whatsapp_links")
def generate_whatsapp_links():
    if 'username' not in session:
        return redirect(url_for('login'))

    print("🚀 بدء إنشاء روابط واتساب...")
    whatsapp_links = check_and_generate_whatsapp_links()

    if whatsapp_links:
        import json
        with open('whatsapp_links.json', 'w', encoding='utf-8') as f:
            json.dump(whatsapp_links, f, ensure_ascii=False, indent=2)

        return redirect(url_for('whatsapp_links_page'))
    else:
        flash("لا يوجد طلاب غائبين اليوم", "info")
        return redirect(url_for('admin'))

@app.route("/whatsapp_links")
def whatsapp_links_page():
    if 'username' not in session:
        return redirect(url_for('login'))

    try:
        import json
        with open('whatsapp_links.json', 'r', encoding='utf-8') as f:
            whatsapp_links = json.load(f)
    except:
        whatsapp_links = []

    return render_template("whatsapp_links.html", links=whatsapp_links, username=session.get('username'))

@app.route("/manual_send_whatsapp/<student_id>")
def manual_send_whatsapp(student_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()
    conn.close()

    if not student_row:
        return redirect(url_for('admin'))

    student_data = dict(student_row)
    phone = student_data['parent_number']
    student_name = student_data['student_name']
    date = today_str()

    clean_phone = ''.join(filter(str.isdigit, str(phone)))
    if not clean_phone.startswith('+'):
        clean_phone = '+2' + clean_phone

    message = f"""تنبيه غياب
عزيزي ولي الأمر،
الطالب/ة {student_name} لم يحضر الحصة اليوم {date}.

يرجى التواصل مع الإدارة للاستفسار عن سبب الغياب.

مع تحيات،
الإدارة""".strip()

    encoded_message = quote(message)
    whatsapp_link = f"https://web.whatsapp.com/send?phone={clean_phone}&text={encoded_message}"

    return redirect(whatsapp_link)

@app.route("/send_monthly_report/<student_id>")
def send_monthly_report(student_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()
    conn.close()

    if not student_row:
        return redirect(url_for('admin'))

    student_data = dict(student_row)
    phone = student_data['parent_number']

    message = generate_detailed_monthly_report_message(student_id)

    if not message:
        flash("لا توجد بيانات لهذا الشهر", "warning")
        return redirect(url_for('admin'))

    clean_phone = ''.join(filter(str.isdigit, str(phone)))
    if not clean_phone.startswith('+'):
        clean_phone = '+2' + clean_phone

    encoded_message = quote(message)
    whatsapp_link = f"https://web.whatsapp.com/send?phone={clean_phone}&text={encoded_message}"

    return redirect(whatsapp_link)

@app.route("/send_present_report/<student_id>")
def send_present_report(student_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT * FROM students WHERE id=?", (student_id,))
    student_row = cursor.fetchone()

    if not student_row:
        conn.close()
        return redirect(url_for('daily_report'))

    student_data = dict(student_row)

    date = today_str()
    cursor = conn.execute("SELECT * FROM history WHERE student_id=? AND date=?", (student_id, date))
    history_row = cursor.fetchone()
    conn.close()

    if history_row:
        today_data = dict(history_row)
        student_data.update(today_data)

    whatsapp_link = generate_present_student_message(student_data)

    if not whatsapp_link:
        flash("تعذر إنشاء رابط واتساب للطالب", "error")
        return redirect(url_for('daily_report'))

    return redirect(whatsapp_link)

@app.route("/download_monthly_reports")
def download_monthly_reports():
    if 'username' not in session:
        return redirect(url_for('login'))

    conn = open_db()
    cursor = conn.execute("SELECT id FROM students")
    student_ids = [row['id'] for row in cursor.fetchall()]
    conn.close()

    from zipfile import ZipFile
    import io

    zip_buffer = io.BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for student_id in student_ids:
            report_path = generate_monthly_report_file(student_id)
            if report_path and os.path.exists(report_path):
                zip_file.write(report_path, os.path.basename(report_path))

    zip_buffer.seek(0)
    return send_file(zip_buffer, download_name=f"monthly_reports_{current_month_str()}.zip", as_attachment=True)

# ---------- Error Handlers ----------
@app.errorhandler(500)
def internal_error(error):
    """التعامل مع أخطاء السيرفر"""
    return render_template('error.html', error=str(error)), 500

@app.errorhandler(404)
def not_found_error(error):
    """التعامل مع صفحات غير موجودة"""
    return render_template('404.html'), 404

@app.errorhandler(ZeroDivisionError)
def handle_zero_division(error):
    """التعامل مع أخطاء القسمة على صفر"""
    flash("حدث خطأ في حساب الإحصائيات. يرجى التحقق من البيانات.", "error")
    return redirect(url_for('admin'))

# ---------- Initialize App ----------
def initialize_app():
    """تهيئة التطبيق عند البدء"""
    init_tables()
    init_db_from_excel()
    mark_absent_for_today()
    print(f"🎯 Server running on PythonAnywhere: https://{PC_IP}")
    print(f"📱 Scanner Page: https://{PC_IP}/remote_scanner")
    print("🔐 نظام تسجيل الدخول مفعل")
    print("👤 المستخدمون المتاحون: admin, teacher")

# تهيئة التطبيق عند الاستيراد
initialize_app()

# لا نستخدم app.run في PythonAnywhere
# PythonAnywhere سيتولى تشغيل التطبيق عبر WSGI